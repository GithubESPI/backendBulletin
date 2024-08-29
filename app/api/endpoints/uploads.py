import base64
import json
import time
import zipfile  # Assurez-vous d'importer le module zipfile standard

# Importation des modules nécessaires
import asyncio
import fitz  # PyMuPDF
from docx2pdf import convert
from fastapi import FastAPI, File, HTTPException, APIRouter, UploadFile
from fastapi.responses import JSONResponse, FileResponse
from openpyxl import load_workbook
from app.core.config import settings
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from docx import Document
import os
import logging
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import requests
from websockets.exceptions import ConnectionClosed

from app.services.api_service import fetch_api_data
from app.services.excel_service import process_excel_file, update_excel_with_appreciations
from app.utils.date_utils import sum_durations

# Configuration du logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

# Création de l'application FastAPI
app = FastAPI()

# Création d'un routeur pour organiser les routes
router = APIRouter()

# Définition du modèle de réponse pour les uploads
class UploadResponse(BaseModel):
    configId: str

# Définition du modèle pour les URLs des documents
class DocumentUrls(BaseModel):
    sessionId: str
    excelUrl: str
    wordUrl: str

# Fonction pour normaliser les titres en supprimant les caractères non alphanumériques et en les mettant en minuscules
def normalize_title(title):
    import re
    if not isinstance(title, str):
        title = str(title)
    return re.sub(r'\W+', '', title).lower()

# Fonction pour récupérer les données d'API en parallèle
async def fetch_api_data_for_template(headers):
    api_urls = [
        f"{settings.YPAERO_BASE_URL}/r/v1/formation-longue/apprenants?codesPeriode=2",
        f"{settings.YPAERO_BASE_URL}/r/v1/formation-longue/groupes",
        f"{settings.YPAERO_BASE_URL}/r/v1/absences/01-01-2023/31-12-2024"
    ]

    api_data_futures = [fetch_api_data(url, headers) for url in api_urls]
    results = await asyncio.gather(*api_data_futures, return_exceptions=True)

    if any(isinstance(result, Exception) for result in results):
        raise HTTPException(status_code=500, detail="Failed to fetch API data")

    return results

# Fonction pour extraire les appréciations depuis un document Word
def extract_appreciations_from_word(word_path):
    try:
        doc = Document(word_path)
        appreciations = {}
        for table in doc.tables:
            for row in table.rows:
                cells = row.cells
                if len(cells) >= 2:
                    name = cells[0].text.strip()
                    appreciation = cells[1].text.strip()
                    if name and appreciation:
                        appreciations[name] = appreciation
        return appreciations
    except Exception as e:
        logger.error("Failed to extract appreciations from Word document", exc_info=True)
    return {}

# Fonction pour logger les données d'un worksheet Excel
def log_excel_data(worksheet):
    data = []
    for row in worksheet.iter_rows(values_only=True):
        data.append(row)
    logger.debug(f"Excel data: {data}")

def extract_code_apprenant(pdf_path: str) -> str:
    try:
        # Ouvrir le document PDF
        with fitz.open(pdf_path) as pdf_document:
            # Parcourir chaque page du document
            for page_num in range(pdf_document.page_count):
                page = pdf_document.load_page(page_num)
                text = page.get_text("text")
                logger.info(f"Extracted text from {pdf_path}: {text}")
                
                # Split le texte en lignes pour trouver la ligne contenant "Identifiant :"
                lines = text.split('\n')
                for line in lines:
                    if "Identifiant :" in line:
                        logger.info(f"Identifiant line found: {line}")
                        parts = line.split("Identifiant :")
                        
                        if len(parts) > 1:
                            code_apprenant = parts[1].strip()
                            logger.info(f"Extracted code_apprenant: {code_apprenant}")
                            
                            if code_apprenant.replace('.', '', 1).isdigit():
                                return str(int(float(code_apprenant)))

                        logger.error(f"No valid code_apprenant found in line: {line}")
                        return None
    
    except Exception as e:
        logger.error(f"Failed to extract code_apprenant from {pdf_path}", exc_info=True)
        return None

def import_document_to_yparéo(file_path, code_apprenant, retries=3, delay=5):
    for attempt in range(retries):
        try:
            with open(file_path, 'rb') as pdf_file:
                file_content = pdf_file.read()
                encoded_content = base64.b64encode(file_content).decode('utf-8')

            # Création du JSON payload pour l'API
            payload = {
                "contenu": encoded_content,
                "nomDocument": os.path.basename(file_path),
                "typeMime": "application/pdf",
                "extension": "pdf",
            }

            # Endpoint Yparéo
            endpoint = f"/r/v1/document/apprenant/{code_apprenant}/document?codeRepertoire=1000011"
            url = f"{settings.YPAERO_BASE_URL}{endpoint}"
            headers = {
                "X-Auth-Token": settings.YPAERO_API_TOKEN,
                "Content-Type": "application/json"
            }

            # Envoi de la requête POST
            response = requests.post(url, headers=headers, json=payload)
            
            if response.status_code == 200:
                return True
            else:
                logging.error(f"Attempt {attempt + 1} failed with status code {response.status_code}: {response.text}")
                if response.status_code == 500:
                    raise ValueError(f"Server error while importing document {file_path}")
        
        except Exception as e:
            logging.error(f"Attempt {attempt + 1} failed due to exception: {str(e)}", exc_info=True)

        time.sleep(delay)
    
    raise ValueError(f"Server error while importing document {file_path} after {retries} retries")

# Fonction pour traiter le fichier téléchargé et intégrer les données dans un template
async def process_file(uploaded_wb, template_path, columns_config, websocket=None):
    try:
        template_wb = openpyxl.load_workbook(template_path, data_only=True)
        uploaded_ws = uploaded_wb.active
        template_ws = template_wb.active

        header_row_uploaded = 4
        header_row_template = 1

        uploaded_titles = {normalize_title(uploaded_ws.cell(row=header_row_uploaded, column=col).value): col
                           for col in range(1, uploaded_ws.max_column + 1) 
                           if uploaded_ws.cell(row=header_row_uploaded, column=col).value is not None}

        template_titles = {normalize_title(template_ws.cell(row=header_row_template, column=col).value): col 
                           for col in range(1, template_ws.max_column + 1) 
                           if template_ws.cell(row=header_row_template, column=col).value is not None}

        matching_columns = {uploaded_title: (uploaded_titles[uploaded_title], template_titles[template_title]) 
                            for uploaded_title in uploaded_titles 
                            for template_title in template_titles 
                            if uploaded_title == template_title}

        if not matching_columns:
            return JSONResponse(content={"message": "No matching columns found, leaving new table empty."})

        template_ws.cell(row=header_row_template + 1, column=columns_config['name_column_index_template']).value = "Nom"

        headers = {
            'X-Auth-Token': settings.YPAERO_API_TOKEN,
            'Content-Type': 'application/json'
        }

        api_data, groupes_data, absences_data = await fetch_api_data_for_template(headers)

        if not isinstance(api_data, dict) or not isinstance(groupes_data, dict) or not isinstance(absences_data, dict):
            raise HTTPException(status_code=500, detail="Unexpected API response format")

        api_dict = {normalize_title(apprenant['nomApprenant'] + apprenant['prenomApprenant']): apprenant for key, apprenant in api_data.items()}
        groupes_dict = {groupe['codeGroupe']: groupe for groupe in groupes_data.values()}
        absences_summary = {}
        for absence in absences_data.values():
            apprenant_id = absence.get('codeApprenant')
            duration = int(absence.get('duree', 0))

            if apprenant_id not in absences_summary:
                absences_summary[apprenant_id] = {'justified': [], 'unjustified': [], 'delays': []}

            if absence.get('isJustifie'):
                absences_summary[apprenant_id]['justified'].append(duration)
            elif absence.get('isRetard'):
                absences_summary[apprenant_id]['delays'].append(duration)
            else:
                absences_summary[apprenant_id]['unjustified'].append(duration)

        exclude_phrase = 'moyennedugroupe'
        total_rows = uploaded_ws.max_row - header_row_uploaded
        processed_rows = 0

        for row in range(header_row_uploaded + 1, uploaded_ws.max_row + 1):
            processed_rows += 1
            progress = (processed_rows / total_rows) * 100

            # Envoyer la progression via WebSocket
            if websocket:
                try:
                    await websocket.send(json.dumps({
                        "progress": progress,
                        "message": f"Processing row {processed_rows} of {total_rows}"
                    }))
                except ConnectionClosed:
                    pass  # Le client a fermé la connexion

            if any(exclude_phrase in normalize_title(uploaded_ws.cell(row=row, column=col).value or '') for col in range(1, uploaded_ws.max_column + 1)):
                continue

            uploaded_name = uploaded_ws.cell(row=row, column=columns_config['name_column_index_uploaded']).value
            template_row = row - header_row_uploaded + header_row_template + 1
            template_ws.cell(row=template_row, column=columns_config['name_column_index_template']).value = uploaded_name

            normalized_name = normalize_title(uploaded_name)

            if (apprenant_info := api_dict.get(normalized_name)):
                template_ws.cell(row=template_row, column=columns_config['code_apprenant_column_index_template']).value = apprenant_info.get('codeApprenant', 'N/A')
                template_ws.cell(row=template_row, column=columns_config['date_naissance_column_index_template']).value = apprenant_info.get('dateNaissance', 'N/A')
                if 'inscriptions' in apprenant_info and apprenant_info['inscriptions']:
                    template_ws.cell(row=template_row, column=columns_config['nom_site_column_index_template']).value = apprenant_info['inscriptions'][0]['site'].get('nomSite', 'N/A')

                code_groupe = apprenant_info.get('informationsCourantes', {}).get('codeGroupe', None)
                if code_groupe and code_groupe in groupes_dict:
                    groupe_info = groupes_dict[code_groupe]
                    template_ws.cell(row=template_row, column=columns_config['code_groupe_column_index_template']).value = groupe_info.get('codeGroupe', 'N/A')
                    template_ws.cell(row=template_row, column=columns_config['nom_groupe_column_index_template']).value = groupe_info.get('nomGroupe', 'N/A')
                    template_ws.cell(row=template_row, column=columns_config['etendu_groupe_column_index_template']).value = groupe_info.get('etenduGroupe', 'N/A')

                apprenant_id = apprenant_info.get('codeApprenant')
                abs_info = absences_summary.get(apprenant_id, {'justified': [], 'unjustified': [], 'delays': []})

                justified_duration = sum_durations(abs_info['justified']) or "00h00"
                unjustified_duration = sum_durations(abs_info['unjustified']) or "00h00"
                delays_duration = sum_durations(abs_info['delays']) or "00h00"

                template_ws.cell(row=template_row, column=columns_config['duree_justifie_column_index_template']).value = justified_duration
                template_ws.cell(row=template_row, column=columns_config['duree_non_justifie_column_index_template']).value = unjustified_duration
                template_ws.cell(row=template_row, column=columns_config['duree_retard_column_index_template']).value = delays_duration

            for uploaded_title, (src_col, dest_col) in matching_columns.items():
                src_cell = uploaded_ws.cell(row=row, column=src_col)
                dest_cell = template_ws.cell(row=template_row, column=dest_col)
                dest_cell.value = src_cell.value

        for col in range(1, template_ws.max_column + 1):
            if template_ws.cell(row=header_row_template + 1, column=col).value == template_ws.cell(row=header_row_template, column=col).value:
                template_ws.cell(row=header_row_template + 1, column=col).value = None

        for col in range(1, template_ws.max_column + 1):
            if template_ws.cell(row=header_row_template + 2, column=col).value == "Note":
                template_ws.cell(row=header_row_template + 2, column=col).value = None

        target_phrase = "* Attention, le total des absences prend en compte toutes les absences aux séances sur la période concernée. S'il existe des absences sur des matières qui ne figurent pas dans le relevé, elles seront également comptabilisées."
        for row in template_ws.iter_rows():
            for cell in row:
                if cell.value == target_phrase:
                    cell.value = None

        log_excel_data(template_ws)
        return template_wb

    except Exception as e:
        logger.error("Failed to process the file", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/upload-and-integrate-excel-and-word")
async def upload_and_integrate(doc_urls: DocumentUrls):
    try:
        # Télécharger le fichier Excel
        excel_response = requests.get(doc_urls.excelUrl)
        if (excel_response.status_code != 200):
            raise HTTPException(status_code=400, detail="Failed to download Excel document")

        temp_excel_path = os.path.join(settings.DOWNLOAD_DIR, f"{doc_urls.sessionId}.xlsx")
        with open(temp_excel_path, 'wb') as temp_excel_file:
            temp_excel_file.write(excel_response.content)

        # Télécharger le fichier Word
        word_response = requests.get(doc_urls.wordUrl)
        if (word_response.status_code != 200):
            raise HTTPException(status_code=400, detail="Failed to download Word document")

        temp_word_path = os.path.join(settings.DOWNLOAD_DIR, f"{doc_urls.sessionId}.docx")
        with open(temp_word_path, 'wb') as temp_word_file:
            temp_word_file.write(word_response.content)

        # Traitement du fichier Excel
        uploaded_wb = load_workbook(temp_excel_path, data_only=True)
        uploaded_ws = uploaded_wb.active

        # Détection du template approprié en fonction des valeurs du fichier Excel
        uploaded_values = [uploaded_ws[cell].value for cell in ['C4', 'F4', 'I4', 'L4', 'O4', 'R4', 'U4', 'X4', 'AA4', 'AD4', 'AG4', 'AJ4', 'AM4', 'AP4', 'AS4', 'AV4', 'AY4', 'BB4', 'BE4', 'BH4'] if uploaded_ws[cell].value is not None]

        templates = {
            "MAPI": settings.M1_S1_MAPI_TEMPLATE,
            "MAGI": settings.M1_S1_MAGI_TEMPLATE,
            "MEFIM": settings.M1_S1_MEFIM_TEMPLATE,
            "MAPI_S2": settings.M1_S2_MAPI_TEMPLATE,
            "MAGI_S2": settings.M1_S2_MAGI_TEMPLATE,
            "MEFIM_S2": settings.M1_S2_MEFIM_TEMPLATE,
            "MAPI_S3": settings.M2_S3_MAPI_TEMPLATE,
            "MAGI_S3": settings.M2_S3_MAGI_TEMPLATE,
            "MEFIM_S3": settings.M2_S3_MEFIM_TEMPLATE,
            "MAPI_S4": settings.M2_S4_MAPI_TEMPLATE,
            "MAGI_S4": settings.M2_S4_MAGI_TEMPLATE,
            "MEFIM_S4": settings.M2_S4_MEFIM_TEMPLATE,
        }

        matching_values = {
            "MAPI": ['UE 1 – Economie & Gestion', 'Stratégie et Solutions Immobilières', 'Finance Immobilière', 'Economie Immobilière I', 'UE 2 – Droit', 'Droit des Affaires et des Contrats', 'UE 3 – Aménagement & Urbanisme', 'Ville et Développements Urbains', "Politique de l'Habitat", 'UE 4 – Compétences Professionnalisantes', 'Real Estate English', "Rencontres de l'Immobilier", 'ESPI Career Services', 'ESPI Inside', 'Immersion Professionnelle', 'Projet Voltaire', 'UE SPE – MAPI', 'Etude Foncière', "Montage d'une Opération de Promotion Immobilière", 'Acquisition et Dissociation du Foncier'],
            "MAGI": ['UE 1 – Economie & Gestion', 'Stratégie et Solutions Immobilières', 'Finance Immobilière', 'Economie Immobilière I', 'UE 2 – Droit', 'Droit des Affaires et des Contrats', 'UE 3 – Aménagement & Urbanisme', 'Ville et Développements Urbains', "Politique de l'Habitat", 'UE 4 – Compétences Professionnalisantes', 'Real Estate English', "Rencontres de l'Immobilier", 'ESPI Career Services', 'ESPI Inside', 'Immersion Professionnelle', 'Projet Voltaire', 'UE SPE – MAGI', 'Baux Commerciaux et Gestion Locative', 'Actifs Tertiaires en Copropriété', 'Techniques du Bâtiment'],
            "MEFIM": ['UE 1 – Economie & Gestion', 'Stratégie et Solutions Immobilières', 'Finance Immobilière', 'Economie Immobilière I', 'UE 2 – Droit', 'Droit des Affaires et des Contrats', 'UE 3 – Aménagement & Urbanisme', 'Ville et Développements Urbains', "Politique de l'Habitat", 'UE 4 – Compétences Professionnalisantes', 'Real Estate English', "Rencontres de l'Immobilier", 'ESPI Career Services', 'ESPI Inside', 'Immersion Professionnelle', 'Projet Voltaire', 'UE SPE – MEFIM', "Les Fondamentaux de l'Evaluation", 'Analyse et Financement Immobilier', 'Modélisation Financière'],
            "MAPI_S2": ['UE 1 – Economie & Gestion', "Marketing de l'Immobilier", 'Investissement et Financiarisation', 'Fiscalité', 'UE 2 – Droit', "Droit de l'Urbanisme et de la Construction", "Déontologie en France et à l'International", 'UE 4 – Compétences Professionnalisantes', 'Immersion Professionnelle', 'Real Estate English', 'Atelier Méthodologie de la Recherche', 'Techniques de Négociation', "Rencontres de l'Immobilier", 'ESPI Inside', 'Projet Voltaire', 'UE SPE – MAPI', "Droit de la Promotion Immobilière", "Montage d'une Opération de Logement", 'Financement des Opérations de Promotion Immobilière', "Logement Social et Accession Sociale"],
            "MAGI_S2": ['UE 1 – Economie & Gestion', "Marketing de l'Immobilier", 'Investissement et Financiarisation', 'Fiscalité', 'UE 2 – Droit', "Droit de l'Urbanisme et de la Construction", "Déontologie en France et à l'International", 'UE 4 – Compétences Professionnalisantes', 'Immersion Professionnelle', 'Real Estate English', 'Atelier Méthodologie de la Recherche', 'Techniques de Négociation', "Rencontres de l'Immobilier", 'ESPI Inside', 'Projet Voltaire', 'UE SPE – MAGI', "Budget d'Exploitation et de Travaux", 'Développement et Stratégie Commerciale', 'Technique et Conformité des Immeubles', "Gestion de l'Immobilier - Logistique et Data Center"],
            "MEFIM_S2": ['UE 1 – Economie & Gestion', "Marketing de l'Immobilier", 'Investissement et Financiarisation', 'Fiscalité', 'UE 2 – Droit', "Droit de l'Urbanisme et de la Construction", "Déontologie en France et à l'International", 'UE 4 – Compétences Professionnalisantes', 'Immersion Professionnelle', 'Real Estate English', 'Atelier Méthodologie de la Recherche', 'Techniques de Négociation', "Rencontres de l'Immobilier", 'ESPI Inside', 'Projet Voltaire', 'UE SPE – MEFIM', "Marché d'Actifs Immobiliers", "Baux Commerciaux", 'Evaluation des Actifs Résidentiels', "Audit et Gestion des Immeubles"],
            "MAPI_S3": ['UE 1 – Economie & Gestion', "PropTech et Innovation", 'Economie Immobilière II', 'UE 3 – Aménagement & Urbanisme', "Stratégies et Aménagement des Territoires I", "UE 4 – Compétences Professionnalisantes", 'Communication Digitale, Ecrite et Orale', 'Immersion Professionnelle', 'Real Estate English', 'Méthodologie de la Recherche', "Rencontres de l'Immobilier", 'ESPI Inside', 'UE SPE – MAPI', "Acquisition et Dissociation du Foncier", "Montage des Opérations Tertiaires", "Aménagement et Commande Publique", "Techniques du Bâtiment", "Réhabilitation et Pathologies du Bâtiment"],
            "MAGI_S3": ['UE 1 – Economie & Gestion', "PropTech et Innovation", 'Economie Immobilière II', 'UE 3 – Aménagement & Urbanisme', "Stratégies et Aménagement des Territoires I", "UE 4 – Compétences Professionnalisantes", 'Communication Digitale, Ecrite et Orale', 'Immersion Professionnelle', 'Real Estate English', 'Méthodologie de la Recherche', "Rencontres de l'Immobilier", 'ESPI Inside', 'UE SPE – MAGI', "Rénovation Energétique des Actifs Tertiaires", "Arbitrage, Optimisation et Valorisation des Actifs Tertiaires", 'Maintenance et Facility Management', "Réhabilitation et Pathologies du Bâtiment"],
            "MEFIM_S3": ['UE 1 – Economie & Gestion', "PropTech et Innovation", 'Economie Immobilière II', 'UE 3 – Aménagement & Urbanisme', "Stratégies et Aménagement des Territoires I", "UE 4 – Compétences Professionnalisantes", 'Communication Digitale, Ecrite et Orale', 'Immersion Professionnelle', 'Real Estate English', 'Méthodologie de la Recherche', "Rencontres de l'Immobilier", 'ESPI Inside', 'UE SPE – MEFIM', "Droit des Suretés et de la Transmission", 'Due Diligence', "Evaluation d'Actifs Tertiaires et Industriels", "Gestion de Patrimoine"],
            "MAPI_S4": ['UE 1 – Economie & Gestion', "Economie de l'Environnement", 'UE 3 – Aménagement & Urbanisme', "Normalisation, Labellisation", "Stratégies et Aménagement des Territoires II", 'UE 4 – Compétences Professionnalisantes', 'Real Estate English', 'Mémoire de Recherche', "Rencontres de l'Immobilier", 'ESPI Career Services', 'Immersion Professionnelle', 'UE SPE – MAPI', "Business Game Aménagement et Promotion Immobilière", "Fiscalité et Promotion Immobilière", "Contentieux de l'Urbanisme"],
            "MAGI_S4": ['UE 1 – Economie & Gestion', "Economie de l'Environnement", 'UE 3 – Aménagement & Urbanisme', "Normalisation, Labellisation", "Stratégies et Aménagement des Territoires II", 'UE 4 – Compétences Professionnalisantes', 'Real Estate English', 'Mémoire de Recherche', "Rencontres de l'Immobilier", 'ESPI Career Services', 'Immersion Professionnelle', 'UE SPE – MAGI', "Business Game Property Management", "Gestion des Centres Commerciaux", "Gestion de Contentieux et Recouvrement"],
            "MEFIM_S4": ['UE 1 – Economie & Gestion', "Economie de l'Environnement", 'UE 3 – Aménagement & Urbanisme', "Normalisation, Labellisation", "Stratégies et Aménagement des Territoires II", 'UE 4 – Compétences Professionnalisantes', 'Real Estate English', 'Mémoire de Recherche', "Rencontres de l'Immobilier", 'ESPI Career Services', 'Immersion Professionnelle', 'UE SPE – MEFIM', "Business Game Arbitrage et Stratégies d'Investissement", "Fiscalité du Patrimoine", "Fintech et Blockchain"],
        }

        column_configs = {
            "MAPI": {
                'name_column_index_uploaded': 2,
                'name_column_index_template': 2,
                'code_apprenant_column_index_template': 1,
                'date_naissance_column_index_template': 23,
                'nom_site_column_index_template': 24,
                'code_groupe_column_index_template': 25,
                'nom_groupe_column_index_template': 26,
                'etendu_groupe_column_index_template': 27,
                'duree_justifie_column_index_template': 28,
                'duree_non_justifie_column_index_template': 29,
                'duree_retard_column_index_template': 30,
                'appreciation_column_index_template': 31
            },
            "MAGI": {
                'name_column_index_uploaded': 2,
                'name_column_index_template': 2,
                'code_apprenant_column_index_template': 1,
                'date_naissance_column_index_template': 23,
                'nom_site_column_index_template': 24,
                'code_groupe_column_index_template': 25,
                'nom_groupe_column_index_template': 26,
                'etendu_groupe_column_index_template': 27,
                'duree_justifie_column_index_template': 28,
                'duree_non_justifie_column_index_template': 29,
                'duree_retard_column_index_template': 30,
                'appreciation_column_index_template': 31
            },
            "MEFIM": {
                'name_column_index_uploaded': 2,
                'name_column_index_template': 2,
                'code_apprenant_column_index_template': 1,
                'date_naissance_column_index_template': 23,
                'nom_site_column_index_template': 24,
                'code_groupe_column_index_template': 25,
                'nom_groupe_column_index_template': 26,
                'etendu_groupe_column_index_template': 27,
                'duree_justifie_column_index_template': 28,
                'duree_non_justifie_column_index_template': 29,
                'duree_retard_column_index_template': 30,
                'appreciation_column_index_template': 31
            },
            "MAPI_S2": {
                'name_column_index_uploaded': 2,
                'name_column_index_template': 2,
                'code_apprenant_column_index_template': 1,
                'date_naissance_column_index_template': 23,
                'nom_site_column_index_template': 24,
                'code_groupe_column_index_template': 25,
                'nom_groupe_column_index_template': 26,
                'etendu_groupe_column_index_template': 27,
                'duree_justifie_column_index_template': 28,
                'duree_non_justifie_column_index_template': 29,
                'duree_retard_column_index_template': 30,
                'appreciation_column_index_template': 31
            },
            "MAGI_S2": {
                'name_column_index_uploaded': 2,
                'name_column_index_template': 2,
                'code_apprenant_column_index_template': 1,
                'date_naissance_column_index_template': 23,
                'nom_site_column_index_template': 24,
                'code_groupe_column_index_template': 25,
                'nom_groupe_column_index_template': 26,
                'etendu_groupe_column_index_template': 27,
                'duree_justifie_column_index_template': 28,
                'duree_non_justifie_column_index_template': 29,
                'duree_retard_column_index_template': 30,
                'appreciation_column_index_template': 31
            },
            "MEFIM_S2": {
                'name_column_index_uploaded': 2,
                'name_column_index_template': 2,
                'code_apprenant_column_index_template': 1,
                'date_naissance_column_index_template': 23,
                'nom_site_column_index_template': 24,
                'code_groupe_column_index_template': 25,
                'nom_groupe_column_index_template': 26,
                'etendu_groupe_column_index_template': 27,
                'duree_justifie_column_index_template': 28,
                'duree_non_justifie_column_index_template': 29,
                'duree_retard_column_index_template': 30,
                'appreciation_column_index_template': 31
            },
            "MAPI_S3": {
                'name_column_index_uploaded': 2,
                'name_column_index_template': 2,
                'code_apprenant_column_index_template': 1,
                'date_naissance_column_index_template': 21,
                'nom_site_column_index_template': 22,
                'code_groupe_column_index_template': 23,
                'nom_groupe_column_index_template': 24,
                'etendu_groupe_column_index_template': 25,
                'duree_justifie_column_index_template': 26,
                'duree_non_justifie_column_index_template': 27,
                'duree_retard_column_index_template': 28,
                'appreciation_column_index_template': 29
            },
            "MAGI_S3": {
                'name_column_index_uploaded': 2,
                'name_column_index_template': 2,
                'code_apprenant_column_index_template': 1,
                'date_naissance_column_index_template': 20,
                'nom_site_column_index_template': 21,
                'code_groupe_column_index_template': 22,
                'nom_groupe_column_index_template': 23,
                'etendu_groupe_column_index_template': 24,
                'duree_justifie_column_index_template': 25,
                'duree_non_justifie_column_index_template': 26,
                'duree_retard_column_index_template': 27,
                'appreciation_column_index_template': 28
            },
            "MEFIM_S3": {
                'name_column_index_uploaded': 2,
                'name_column_index_template': 2,
                'code_apprenant_column_index_template': 1,
                'date_naissance_column_index_template': 20,
                'nom_site_column_index_template': 21,
                'code_groupe_column_index_template': 22,
                'nom_groupe_column_index_template': 23,
                'etendu_groupe_column_index_template': 24,
                'duree_justifie_column_index_template': 25,
                'duree_non_justifie_column_index_template': 26,
                'duree_retard_column_index_template': 27,
                'appreciation_column_index_template': 28
            },
            "MAPI_S4": {
                'name_column_index_uploaded': 2,
                'name_column_index_template': 2,
                'code_apprenant_column_index_template': 1,
                'date_naissance_column_index_template': 18,
                'nom_site_column_index_template': 19,
                'code_groupe_column_index_template': 20,
                'nom_groupe_column_index_template': 21,
                'etendu_groupe_column_index_template': 22,
                'duree_justifie_column_index_template': 23,
                'duree_non_justifie_column_index_template': 24,
                'duree_retard_column_index_template': 25,
                'appreciation_column_index_template': 26
            },
            "MAGI_S4": {
                'name_column_index_uploaded': 2,
                'name_column_index_template': 2,
                'code_apprenant_column_index_template': 1,
                'date_naissance_column_index_template': 18,
                'nom_site_column_index_template': 19,
                'code_groupe_column_index_template': 20,
                'nom_groupe_column_index_template': 21,
                'etendu_groupe_column_index_template': 22,
                'duree_justifie_column_index_template': 23,
                'duree_non_justifie_column_index_template': 24,
                'duree_retard_column_index_template': 25,
                'appreciation_column_index_template': 26
            },
            "MEFIM_S4": {
                'name_column_index_uploaded': 2,
                'name_column_index_template': 2,
                'code_apprenant_column_index_template': 1,
                'date_naissance_column_index_template': 18,
                'nom_site_column_index_template': 19,
                'code_groupe_column_index_template': 20,
                'nom_groupe_column_index_template': 21,
                'etendu_groupe_column_index_template': 22,
                'duree_justifie_column_index_template': 23,
                'duree_non_justifie_column_index_template': 24,
                'duree_retard_column_index_template': 25,
                'appreciation_column_index_template': 26
            },
        }

        template_to_use = None
        columns_config = None
        for template, values in matching_values.items():
            if uploaded_values[:len(values)] == values:
                template_to_use = templates[template]
                columns_config = column_configs[template]
                break

        if not template_to_use or not columns_config:
            raise HTTPException(status_code=400, detail="No matching template found")

        # Processer le fichier et créer le fichier Excel final
        template_wb = await process_file(uploaded_wb, template_to_use, columns_config)

        appreciations = extract_appreciations_from_word(temp_word_path)
        logger.debug(f"Extracted appreciations: {appreciations}")

        template_wb = update_excel_with_appreciations(template_wb, appreciations, columns_config)

        template_name = os.path.basename(template_to_use).replace('.xlsx', '')
        output_path = os.path.join(settings.OUTPUT_DIR, f'{template_name}.xlsx')
        template_wb.save(output_path)

        # Génération et importation des bulletins PDF
        output_dir = os.path.join(settings.OUTPUT_DIR, 'bulletins')
        os.makedirs(output_dir, exist_ok=True)
        bulletin_paths = process_excel_file(output_path, output_dir)
        convert(output_dir)

        pdf_bulletin_paths = [
            os.path.join(output_dir, filename.replace('.docx', '.pdf'))
            for filename in os.listdir(output_dir)
            if filename.endswith('.pdf')
        ]

        import_errors = []
        for pdf_bulletin_path in pdf_bulletin_paths:
            try:
                # Ajouter un log avant de traiter chaque bulletin
                logger.info(f"Processing bulletin for apprenant {pdf_bulletin_path}")
                
                code_apprenant = extract_code_apprenant(pdf_bulletin_path)
                if not code_apprenant:
                    raise ValueError(f"codeApprenant not found in {pdf_bulletin_path}")
                
                logger.info(f"Extracted codeApprenant: {code_apprenant} from {pdf_bulletin_path}")

                if not import_document_to_yparéo(pdf_bulletin_path, code_apprenant):
                    raise ValueError(f"Failed to import document {pdf_bulletin_path}")

            except Exception as import_exc:
                logger.error(f"Failed to import bulletin {pdf_bulletin_path}", exc_info=True)
                import_errors.append({
                    "file": os.path.basename(pdf_bulletin_path),
                    "error": str(import_exc)
                })

        # Création d'un fichier ZIP avec les PDF générés
        zip_filename = os.path.join(settings.DOWNLOAD_DIR, 'bulletins.zip')
        with zipfile.ZipFile(zip_filename, 'w') as zipf:
            for pdf_path in pdf_bulletin_paths:
                zipf.write(pdf_path, os.path.basename(pdf_path))

        for bulletin_path in bulletin_paths:
            os.remove(bulletin_path)

        if import_errors:
            return JSONResponse(content={"message": "Bulletins PDF generated, but some failed to import", "errors": import_errors}, status_code=207)
        else:
            return JSONResponse(content={"message": "Bulletins PDF generated, imported to Yparéo, and zipped successfully", "zip_path": zip_filename})
    
    except Exception as e:
        logger.error("Failed to process the file and generate bulletins", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/generate-and-import-bulletins")
async def generate_and_import_bulletins(file: UploadFile = File(...)):
    try:
        # Step 1: Save the uploaded file
        temp_file_path = os.path.join(settings.UPLOAD_DIR, file.filename)
        with open(temp_file_path, 'wb') as temp_file:
            temp_file.write(await file.read())

        # Step 2: Generate PDF bulletins from the Excel file
        output_dir = os.path.join(settings.OUTPUT_DIR, 'bulletins')
        os.makedirs(output_dir, exist_ok=True)
        bulletin_paths = process_excel_file(temp_file_path, output_dir)

        # Convert all Word documents to PDF (Assuming process_excel_file generates Word docs)
        convert(output_dir)
        pdf_bulletin_paths = [
            os.path.join(output_dir, filename.replace('.docx', '.pdf'))
            for filename in os.listdir(output_dir)
            if filename.endswith('.pdf')
        ]

        # Step 3: Import each PDF bulletin to Yparéo
        import_errors = []
        for pdf_bulletin_path in pdf_bulletin_paths:
            try:
                # Extract codeApprenant from the PDF
                code_apprenant = extract_code_apprenant(pdf_bulletin_path)
                if not code_apprenant:
                    raise ValueError(f"codeApprenant not found in {pdf_bulletin_path}")
                
                logger.info(f"Extracted codeApprenant: {code_apprenant} from {pdf_bulletin_path}")

                if not import_document_to_yparéo(pdf_bulletin_path, code_apprenant):
                    raise ValueError(f"Failed to import document {pdf_bulletin_path}")

            except Exception as import_exc:
                logger.error(f"Failed to import bulletin {pdf_bulletin_path}", exc_info=True)
                import_errors.append({
                    "file": os.path.basename(pdf_bulletin_path),
                    "error": str(import_exc)
                })

        # Step 4: Create a ZIP file with the PDF bulletins after import to Yparéo
        zip_filename = os.path.join(settings.DOWNLOAD_DIR, 'bulletins.zip')
        with zipfile.ZipFile(zip_filename, 'w') as zipf:
            for pdf_path in pdf_bulletin_paths:
                zipf.write(pdf_path, os.path.basename(pdf_path))

        for bulletin_path in bulletin_paths:
            os.remove(bulletin_path)

        if import_errors:
            return JSONResponse(content={"message": "Bulletins PDF generated, but some failed to import", "errors": import_errors}, status_code=207)
        else:
            return JSONResponse(content={"message": "Bulletins PDF generated, imported to Yparéo, and zipped successfully", "zip_path": zip_filename})

    except Exception as e:
        logger.error("Failed to generate and import bulletins", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/download-zip/{filename}")
async def download_zip(filename: str):
    zip_path = os.path.join(settings.DOWNLOAD_DIR, filename)
    if not os.path.exists(zip_path):
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(path=zip_path, filename=filename, media_type='application/zip')